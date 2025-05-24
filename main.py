import openpyxl

# Solicitar nombres y notas de los estudiantes
estudiantes = {}
for i in range(3):
    nombre = input(f"Ingrese el nombre del estudiante {i + 1}: ")
    notas = []
    for j in range(3):  # Suponiendo que cada estudiante tiene 3 notas
        nota = float(input(f"Ingrese la nota {j + 1} de {nombre}: "))
        notas.append(nota)
    estudiantes[nombre] = notas

# Calcular el promedio de las notas de cada estudiante
promedios = {nombre: sum(notas) / len(notas) for nombre, notas in estudiantes.items()}

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active

# Escribir los nombres en la columna A y los promedios en la columna B
fila = 1
for nombre, promedio in promedios.items():
    ws[f"A{fila}"] = nombre
    ws[f"B{fila}"] = promedio
    fila += 1

# Escribir el promedio general en la celda B1
promedio_general = sum(promedios.values()) / len(promedios)
ws["B1"] = promedio_general

# Guardar el archivo
wb.save("estudiantes_promedios.xlsx")

print("Los datos han sido guardados en el archivo 'estudiantes_promedios.xlsx'.")
